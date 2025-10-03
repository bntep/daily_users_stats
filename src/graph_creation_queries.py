import os
import pandas as pd
import sys
import argparse
from pathlib import Path
import seaborn as sns
import matplotlib.pyplot as plt
sys.path.append(str(Path(os.getcwd())))
from utils.dbclient.DatabaseClient import DbConnector
from openpyxl import load_workbook
from module.env import *

#DEV_python = Path("C:/Users/akash/Documents/DEV_python")

condition_year= f"date_part('year',date_heure_extraction) IN (2021,2022,2023,2024,2025)" 
condition_labo = ""

# Define base folder
result_dir = Path("/home/groups/daily/travail/Bertrand/Developpement/daily_users_stats/stat_graphs_files")

def create_statistique_requete(condition_year: str, type_data: str , condition_labo: str = "") -> pd.DataFrame:
    req_extraction_stats = f"SELECT distinct id_utilisateur_drupal as id_user, nom_utilisateur as user_name ,id_groupe_labo as id_labo , node.name as institution_name, \
    date_part('year',date_heure_extraction) as year, date_part('month',date_heure_extraction) as month, nom_base_interrogee as database_name, \
    type_interrogation, nb_codes_en_entree as nb_codes, date_heure_extraction,CASE date_part('month',date_heure_extraction) \
    WHEN 1 THEN 'January' \
    WHEN 2 THEN 'February' \
    WHEN 3 THEN 'March' \
    WHEN 4 THEN 'April' \
    WHEN 5 THEN 'May' \
    WHEN 6 THEN 'June' \
    WHEN 7 THEN 'July' \
    WHEN 8 THEN 'August' \
    WHEN 9 THEN 'September' \
    WHEN 10 THEN 'October' \
    WHEN 11 THEN 'November' \
    WHEN 12 THEN 'December' \
    END as month2, \
    CASE type_interrogation \
    WHEN 1 THEN 'prévisualisation' \
    WHEN 2 THEN 'téléchargement' \
    WHEN 3 THEN 'téléchargement' \
    END as type_interrogation2, \
    CASE \
    WHEN nom_base_interrogee LIKE '%histo_actions%' THEN 'Stocks' \
    WHEN nom_base_interrogee LIKE '%actions_%' THEN 'Stocks' \
    WHEN nom_base_interrogee LIKE '%indices_telekurs%' THEN 'Global\Market Indices' \
    WHEN nom_base_interrogee LIKE '%histo_indices_telekurs%' THEN 'Global\Market Indices' \
    WHEN nom_base_interrogee LIKE '%indices_eurofidai%' THEN 'Eurofidai Indices' \
    WHEN nom_base_interrogee LIKE '%histo_indices_eurofidai%' THEN 'IEurofidai Indices' \
    WHEN nom_base_interrogee LIKE '%corres_code%' THEN 'Code Mapping Table' \
    WHEN nom_base_interrogee LIKE '%fonds_mutuel_%' THEN 'Mutual Funds' \
    WHEN nom_base_interrogee LIKE '%change%' THEN 'Spot Exchange Rate' \
    WHEN nom_base_interrogee LIKE '%histo_ost%' THEN 'Corporate Events' \
    WHEN nom_base_interrogee LIKE 'ost%' THEN 'Corporate Events' \
    WHEN nom_base_interrogee LIKE '%esg%' THEN 'ESG' \
    WHEN nom_base_interrogee LIKE '%greenbonds%' THEN 'Green Bonds' \
    END as database_name2, \
    CASE \
    WHEN nom_base_interrogee LIKE '%histo%' THEN 'Search_Code' \
    ELSE 'Extract_Data' \
    END AS code_ou_data \
    FROM statistique_requete as sr LEFT JOIN institution_entity as node \
    ON sr.id_groupe_labo=node.id \
    WHERE {condition_year} AND  nom_groupe_labo NOT IN ('EUROFIDAI','administrateur Drupal') \
    AND id_utilisateur_drupal NOT IN (1178,1922,367) {condition_labo} \
    ORDER BY year,month,date_heure_extraction,id_utilisateur_drupal,node.name \
    ;"

    req_extraction_stats_users="select distinct ufd.uid as id_user, ie.name as labo_name, to_timestamp(ufd.created)::date AS date_created, \
        to_timestamp(access)::date AS date_last_access, ufs.field_statut_value AS statut \
        FROM users_field_data AS ufd \
            LEFT OUTER JOIN user__roles AS ur ON ur.entity_id=ufd.uid \
            LEFT OUTER JOIN user__field_institution AS ufi ON ufi.entity_id=ufd.uid \
            LEFT OUTER JOIN institution_entity AS ie ON ie.id=field_institution_target_id \
            FULL OUTER JOIN user__field_statut AS ufs  ON ufd.uid =ufs.entity_id \
            WHERE ufd.uid NOT IN (1178,1922,367) AND ie.name NOT IN ('EUROFIDAI','administrateur Drupal')  \
            ORDER BY  ie.name,  id_user ;"

    #DEV_python = Path("C:/Users/akash/Documents/DEV_python")

    # condition_year = f"date_part('year',date_heure_extraction) IN (2021,2022,2023,2024,2025)" 
    # condition_labo = ""

    # Your SQL queries unchanged...
    # req_extraction_stats and req_extraction_stats_users defined above

    # Execute queries
    df_stats_daily_users = DbConnector('yakari', echo=True).execute_query(req_extraction_stats)
    df_stats_daily_subscription = DbConnector('yakari', echo=True).execute_query(req_extraction_stats_users)

    # Ensure types
    df_stats_daily_users['year'] = df_stats_daily_users['year'].astype(int)
    df_stats_daily_users['month'] = df_stats_daily_users['month'].astype(int)
    df_stats_daily_users['user_full'] = df_stats_daily_users['user_name'] + ' - ' + df_stats_daily_users['institution_name']
    df_stats_daily_users['month_full'] = df_stats_daily_users['month'].astype(str).str.zfill(2) + ' - ' + df_stats_daily_users['month2']

    df_stats_daily_users['user_full'] = (
        df_stats_daily_users['user_name'].fillna('Unknown User').astype(str) +
        ' - ' +
        df_stats_daily_users['institution_name'].fillna('Unknown Institution').astype(str)
    )


    # Merge user status info
    df_stats_daily_users = pd.merge(
        df_stats_daily_users,
        df_stats_daily_subscription,
        how='left',
        left_on='id_user',
        right_on='id_user'
    )
    return df_stats_daily_users



def update_excel_sheet_with_dataframe(file_path, sheet_name, dataframe):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Clear previous data (excluding header row)
    ws.delete_rows(2, ws.max_row)

    # Write new data starting from row 2
    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(file_path)
    print(f"Updated: {file_path} (sheet: {sheet_name})")


def create_excel_statistics(data: pd.DataFrame):

    #create directory if not exists
    result_dir.mkdir(parents=True, exist_ok=True)
    
    # --- Save number of unique users ---   
    df_stats_number_of_unique_users = data.groupby(['year','month_full','month2'])['user_name'].nunique().reset_index(name='nb_unique_users')
    df_stats_number_of_unique_users.sort_values(by=['year','month_full','month2'], inplace=True)
    
    # 1. Unique Users
    update_excel_sheet_with_dataframe(
        file_path=result_dir / "stats_number_of_unique_users.xlsx",
        sheet_name="UniqueUsers",
        dataframe=df_stats_number_of_unique_users
    )

    # --- Save sum of codes per month ---
    df_stats_sum_of_codes = data.groupby(['year', 'month_full','month2'])['nb_codes'].sum().reset_index(name='sum_of_codes')
    df_stats_sum_of_codes.sort_values(by=['year', 'month_full','month2'], inplace=True)
    
    # 2. Sum of Codes
    update_excel_sheet_with_dataframe(
        file_path=result_dir / "stats_sum_of_codes.xlsx",
        sheet_name="TotalCodes",
        dataframe=df_stats_sum_of_codes
    )

    # --- Save user-level summary ---
    df_user_monthly_sum = data.groupby(['user_full', 'statut', 'database_name', 'year', 'month_full','month2'])['nb_codes'].sum().reset_index(name='sum_of_codes')
    df_user_monthly_sum.sort_values(by=['user_full', 'year', 'month_full','month2'], inplace=True)
    # 3. User Monthly Sum
    update_excel_sheet_with_dataframe(
        file_path=result_dir / "user_monthly_sum_of_codes.xlsx",
        sheet_name="Users",
        dataframe=df_user_monthly_sum
    )

    # --- Save institution-level summary ---
    df_institution_monthly_sum = data.groupby(['institution_name', 'database_name', 'year', 'month_full','month2'])['nb_codes'].sum().reset_index(name='sum_of_codes')
    df_institution_monthly_sum.sort_values(by=['institution_name', 'year', 'month_full','month2'], inplace=True)
    # 4. Institution Monthly Sum
    update_excel_sheet_with_dataframe(
        file_path=result_dir / "institution_monthly_sum_of_codes.xlsx",
        sheet_name="Institutions",
        dataframe=df_institution_monthly_sum
    )


if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Generate statistics and graphs from user data.")
    parser.add_argument('--year', type=str, default="2021,2022,2023,2024,2025", help="Years to include in the format 'YYYY,YYYY,...'")
    parser.add_argument('--labo', type=str, default="", help="Laboratory condition for SQL query (e.g., AND id_groupe_labo=3)")
    args = parser.parse_args()

    condition_year = f"date_part('year',date_heure_extraction) IN ({args.year})"
    condition_labo = args.labo

    # Fetch data
    df_stats_daily_users = create_statistique_requete(condition_year, "all", condition_labo)

    # Create Excel statistics
    create_excel_statistics(df_stats_daily_users)